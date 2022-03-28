import os.path
import datetime
import openpyxl

dt_now = datetime.datetime.now()
dir_path = "D:/vocab"
file_path = f"{dir_path}/{dt_now.date()}.xlsx"

def save_excel(word, summary):
    if word == "" or summary == "":
        return
    if os.path.exists(dir_path) == False:
        os.mkdir(dir_path)
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1, value='word')
        sheet.cell(row=1, column=2, value='mean')
    last_row = sheet.max_row + 1
    sheet.cell(row=last_row, column=1, value=word)
    sheet.cell(row=last_row, column=2, value=summary)
    wb.save(file_path)